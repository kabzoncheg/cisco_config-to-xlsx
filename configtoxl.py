import argparse
import os
import re

import openpyxl
from ciscoconfparse import CiscoConfParse


def get_args():
    """
    get_args() gets arguments from user input, when run from cmd
    :return: dict with arguments
    """
    parser = argparse.ArgumentParser(description=''
                                                 'Script for parsing .txt files populated with various "show" commands'
                                                 'output from Cisco network devices and populating .xls document'
                                                 'with required information: IP-addresses, S/N, software version, etc.')
    parser.add_argument('-X', '--xls', type=str, help='Path to .xls document',
                        default=os.path.join(os.getcwd(), 'xls_template.xlsx'))
    parser.add_argument('-S', '--show', type=str, help='Path to folder with .txt files', default=os.getcwd())

    args = parser.parse_args()
    if not args.xls:
        raise ValueError('Path to .xls document must be specified')
    return args


def show_cmd_parser(path_to_file):
    """
    func show_cmd_parser
    :param path_to_file: OS Path to file
    :return: dict with device parameters

    It is better to provide "show tech-support" for this parser
    For device atributes such as hostname, model, serial number etc simple regex are used
    For interface information parsing ciscoconfparse module is used
    """
    parsed_values = {}
    dict_for_ciscoconfparser = []
    with open(path_to_file) as open_file:

        # This cycle gets info from 'show version' output with simple regex
        for line in open_file:
            dict_for_ciscoconfparser.append(line)
            if re.match('^Model number*', line):
                value = (line.rstrip()).split()[-1]
                parsed_values['model'] = value
                parsed_values['device_type'] = 'SWITCH'

            if re.match('^.+ bytes of memory', line):
                value = re.search('\S+ \(.+?\)', line)
                parsed_values['model'] = value.group(0)
                parsed_values['device_type'] = 'ROUTER'

            if re.match('^hostname *', line):
                value = (line.rstrip()).split()[-1]
                parsed_values['hostname'] = value

            if re.match('^System serial number *', line):
                value = (line.rstrip()).split()[-1]
                parsed_values['serial_number'] = value

            if re.match('^Processor board ID *', line):
                value = (line.rstrip()).split()[-1]
                parsed_values['serial_number'] = value

            if re.match('.+ Software .+, Version', line):
                lic = re.search('\(.+?\)', line)
                parsed_values['software_license'] = lic.group(0).strip('( , )')
                softver = re.search(' [0-9]+\.\S+', line)
                parsed_values['software_version'] = softver.group(0).strip(',')

    parse = CiscoConfParse(dict_for_ciscoconfparser, factory=True)
    interface_object_list = parse.find_objects('^interface\s')
    parsed_values['interfaces'] = {}

    # This cycle gets information from 'show run' output
    for obj in interface_object_list:
        parsed_values['interfaces'][obj.name] = {}
        parsed_values['interfaces'][obj.name]['int_name'] = obj.name
        parsed_values['interfaces'][obj.name]['int_description'] = obj.description
        parsed_values['interfaces'][obj.name]['int_ip_address'] = obj.ip_addr

        if obj.is_switchport is True:
            parsed_values['interfaces'][obj.name]['port_type'] = 'switchport'
            parsed_values['interfaces'][obj.name]['int_vlan'] = obj.access_vlan
            parsed_values['interfaces'][obj.name]['int_trunk'] = obj.has_manual_switch_trunk
        else:
            parsed_values['interfaces'][obj.name]['port_type'] = None

    # This cycles get up/down information form 'show interfaces' output
    for interface in parsed_values['interfaces']:
        for element in dict_for_ciscoconfparser:
            regex_for_match = interface + r' is ((up)|(administratively down)|(down)), line protocol is ((up)|(down))'
            if re.match(regex_for_match, element):
                if re.match('^.+ line protocol is up', element):
                    parsed_values['interfaces'][interface]['int_status'] = 'UP'
                    break
                if re.match('^.+ line protocol is down', element):
                    parsed_values['interfaces'][interface]['int_status'] = 'DOWN'
                    break
            else:
                parsed_values['interfaces'][interface]['int_status'] = 'NOT CONFIGURED'
    return parsed_values


def xlsx_writer(dict_list, xlsx_file):
    """
    xlsx_writeer is function for writing from dictionary to .xlsx files
    :param dict_list: list of parsed dictionarys with devices parameters
    :param xlsx_file: path to .xlsx file
    :return: None
    This function determines sheet in .xlsx document automaticaly, based on dict['device_type] value
    """
    wb = openpyxl.load_workbook(xlsx_file)

    for cmd_dict in dict_list:

        sheet_name = cmd_dict['device_type']
        try:
            sheet = wb.get_sheet_by_name(sheet_name)
        except:
            error_msg = 'Cannot write to {}, sheet {} does not exist!'.format(xlsx_file, sheet_name)
            raise ValueError(error_msg)

        # 'F' Column is the longest. So we have to check for the available space
        sheet_offset = 2
        while sheet.cell(row=sheet_offset, column=6).value is not None:
            sheet_offset += 1

        sheet.cell(row=sheet_offset, column=1).value = cmd_dict['hostname']
        sheet.cell(row=sheet_offset, column=2).value = cmd_dict['model']
        sheet.cell(row=sheet_offset, column=3).value = cmd_dict['software_version']
        sheet.cell(row=sheet_offset, column=4).value = cmd_dict['software_license']
        sheet.cell(row=sheet_offset, column=5).value = cmd_dict['serial_number']
        int_list = list(interface for interface in cmd_dict['interfaces'])
        int_list.sort()
        for element in int_list:
            sheet.cell(row=sheet_offset, column=6).value = cmd_dict['interfaces'][element]['int_name']
            sheet.cell(row=sheet_offset, column=7).value = cmd_dict['interfaces'][element]['int_ip_address']
            sheet.cell(row=sheet_offset, column=8).value = cmd_dict['interfaces'][element]['int_status']
            sheet.cell(row=sheet_offset, column=9).value = cmd_dict['interfaces'][element]['int_description']
            if cmd_dict['interfaces'][element]['port_type'] == 'switchport':
                sheet.cell(row=sheet_offset, column=10).value = cmd_dict['interfaces'][element]['int_vlan']
                sheet.cell(row=sheet_offset, column=11).value = cmd_dict['interfaces'][element]['int_trunk']
            sheet_offset += 1

    wb.save(xlsx_file)


def main():
    parsed_output = []
    args = get_args()
    path_to_folder = args.show
    path_to_xls_file = args.xls

    txt_files_paths = [os.path.join(path_to_folder, file)
                       for file in os.listdir(path_to_folder)
                       if file.endswith('.txt') and os.path.isfile(file)]

    for file in txt_files_paths:
        parsed_output.append(show_cmd_parser(file))
    xlsx_writer(parsed_output, path_to_xls_file)


if __name__ == main():
    main()
